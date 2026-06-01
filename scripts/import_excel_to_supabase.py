#!/usr/bin/env python3
"""Seed or sync Stack n Stock Excel data into Supabase crm_accounts.

Usage:
  python scripts/import_excel_to_supabase.py --excel "C:\\Users\\Abhinand\\Downloads\\SNS_CRM.xlsx"

Required env vars:
  SUPABASE_URL
  SUPABASE_SERVICE_ROLE_KEY (preferred) or SUPABASE_ANON_KEY
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple
from urllib import error, parse, request

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    print("openpyxl is required. Use the bundled runtime Python or install openpyxl.", file=sys.stderr)
    raise


ACTION_PROBABILITY = {
    "Approach now": 60,
    "Shortlist": 40,
    "Validate": 25,
    "Approach week later": 20,
}

STAGE_BY_STATUS = {
    "Not Started": "Prospecting",
    "Researching": "Qualified",
    "Email Sent": "Qualified",
    "Meeting Booked": "Proposal",
    "In Discussion": "Proposal",
    "Partner Confirmed": "Won",
    "On Hold": "Lost",
}

PRIORITY_BY_TIER = {
    "Tier 1": "P1",
    "Tier 2": "P2",
    "Strategic": "P0",
}

SCORE_WEIGHTS = {
    "score_qc_urgency": 15,
    "score_sku_fit": 20,
    "score_order_density": 15,
    "score_pilot_willing": 15,
    "score_accessibility": 15,
    "score_logo_value": 20,
}


def cli() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import SNS_CRM.xlsx into Supabase crm_accounts")
    parser.add_argument(
        "--excel",
        default=r"C:\Users\Abhinand\Downloads\SNS_CRM.xlsx",
        help="Path to SNS_CRM.xlsx",
    )
    parser.add_argument("--table", default="crm_accounts", help="Destination accounts table")
    parser.add_argument("--channels-table", default="crm_channels", help="Destination channels table")
    parser.add_argument("--competitors-table", default="crm_competitors", help="Destination competitors table")
    parser.add_argument("--batch-size", type=int, default=200, help="Upsert batch size")
    parser.add_argument("--reset", action="store_true", help="Delete all rows in account table before upsert")
    parser.add_argument("--reset-competitors", action="store_true", help="Delete all rows in competitors table before upsert")
    parser.add_argument("--dry-run", action="store_true", help="Print payload summary only")
    return parser.parse_args()


def parse_value(v: Any) -> Any:
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return v


def parse_date(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, date):
        return v.isoformat()

    raw = str(v).strip()
    if not raw:
        return None

    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%a %b %d %Y %H:%M:%S GMT%z"):
        try:
            dt = datetime.strptime(raw, fmt)
            return dt.date().isoformat()
        except ValueError:
            continue

    try:
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        return dt.date().isoformat()
    except ValueError:
        return None


def parse_datetime(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day).isoformat()

    raw = str(v).strip()
    if not raw:
        return None

    for fmt in (
        "%Y-%m-%dT%H:%M:%S.%fZ",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%a %b %d %Y %H:%M:%S GMT%z",
    ):
        try:
            dt = datetime.strptime(raw, fmt)
            return dt.isoformat()
        except ValueError:
            continue

    try:
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        return dt.isoformat()
    except ValueError:
        return None


def split_platforms(raw: Any) -> List[str]:
    if raw is None:
        return []
    txt = str(raw).strip()
    if not txt:
        return []

    if "|" in txt:
        parts = [p.strip() for p in txt.split("|") if p.strip()]
    else:
        parts = [p.strip() for p in txt.split(",") if p.strip()]

    dedup: List[str] = []
    seen = set()
    for p in parts:
        key = p.lower()
        if key in seen:
            continue
        seen.add(key)
        dedup.append(p)
    return dedup


def channels_from_platforms(platforms: List[str]) -> List[Dict[str, Any]]:
    if not platforms:
        return []
    base = 100 // len(platforms)
    remainder = 100 - base * len(platforms)
    out = []
    for idx, name in enumerate(platforms):
        share = base + (remainder if idx == 0 else 0)
        out.append({"name": name, "pct_share": share})
    return out


def channel_rows_from_accounts(accounts: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    names = set()
    for row in accounts:
        for ch in row.get("channels", []):
            if ch.get("name"):
                names.add(ch["name"].strip())

    rows = []
    for name in sorted(names):
        lower = name.lower()
        ctype = "q-commerce"
        if lower in {"amazon", "flipkart", "nykaa"}:
            ctype = "marketplace"
        elif lower in {"d2c", "own store", "own site", "website"}:
            ctype = "direct"

        rows.append(
            {
                "name": name,
                "type": ctype,
                "notes": "Seeded from SNS_CRM.xlsx import",
            }
        )
    return rows


def calc_qc_score(row: Dict[str, Any]) -> int:
    total = (
        int(row["score_qc_urgency"]) * SCORE_WEIGHTS["score_qc_urgency"]
        + int(row["score_sku_fit"]) * SCORE_WEIGHTS["score_sku_fit"]
        + int(row["score_order_density"]) * SCORE_WEIGHTS["score_order_density"]
        + int(row["score_pilot_willing"]) * SCORE_WEIGHTS["score_pilot_willing"]
        + int(row["score_accessibility"]) * SCORE_WEIGHTS["score_accessibility"]
        + int(row["score_logo_value"]) * SCORE_WEIGHTS["score_logo_value"]
    )
    return round(total / 5)


def clamp_rating(v: Any) -> int:
    try:
        n = int(float(v))
    except (TypeError, ValueError):
        n = 3
    if n < 1:
        return 1
    if n > 5:
        return 5
    return n


def as_int(v: Any, default: int = 0) -> int:
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def normalize_action(v: Any) -> str:
    raw = str(v or "").strip()
    if not raw:
        return "Shortlist"
    if raw.lower() == "approach week later":
        return "Approach week later"
    if raw.lower() == "approach now":
        return "Approach now"
    if raw.lower() == "shortlist":
        return "Shortlist"
    if raw.lower() == "validate":
        return "Validate"
    return "Shortlist"


def build_contact_map(wb) -> Dict[str, Dict[str, Any]]:
    if "Contacts" not in wb.sheetnames:
        return {}

    ws = wb["Contacts"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}

    def get(r: int, h: str) -> Any:
        if h not in idx:
            return None
        return parse_value(ws.cell(r, idx[h]).value)

    out: Dict[str, Dict[str, Any]] = {}
    for r in range(2, ws.max_row + 1):
        account_id = str(get(r, "accountId") or "").strip()
        if not account_id:
            continue

        item = {
            "name": str(get(r, "name") or "").strip(),
            "designation": str(get(r, "designation") or "").strip(),
            "email": str(get(r, "email") or "").strip(),
            "phone": str(get(r, "phone") or "").strip(),
            "linkedin": str(get(r, "linkedin") or "").strip(),
            "nextstep": str(get(r, "nextstep") or "").strip(),
        }

        has_data = any(item[k] for k in ("name", "designation", "email", "phone", "linkedin", "nextstep"))
        if not has_data:
            continue

        if account_id not in out:
            out[account_id] = item
    return out


def build_accounts_payload(path: str) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    if "Accounts" not in wb.sheetnames:
        raise ValueError("Accounts sheet not found in workbook")

    contacts = build_contact_map(wb)
    ws = wb["Accounts"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}

    def get(r: int, h: str) -> Any:
        if h not in idx:
            return None
        return parse_value(ws.cell(r, idx[h]).value)

    rows: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        legacy_id = str(get(r, "id") or "").strip()
        name = str(get(r, "name") or "").strip()
        if not legacy_id or not name:
            continue

        action = normalize_action(get(r, "action"))
        probability = ACTION_PROBABILITY.get(action, 40)
        stage = STAGE_BY_STATUS.get(str(get(r, "status") or "").strip(), "Prospecting")
        priority = PRIORITY_BY_TIER.get(str(get(r, "tier") or "").strip(), "P2")

        platforms = split_platforms(get(r, "platforms"))
        channels = channels_from_platforms(platforms)

        contact = contacts.get(legacy_id, {})

        row = {
            "legacy_id": legacy_id,
            "name": name,
            "company_type": str(get(r, "category") or "").strip(),
            "contact_name": contact.get("name") or "",
            "contact_email": contact.get("email") or "",
            "contact_phone": contact.get("phone") or "",
            "owner": "",
            "city": "Bangalore",
            "stage": stage,
            "deal_value": 0,
            "probability": probability,
            "score": as_int(get(r, "score"), 0),
            "next_action_at": None,
            "next_action": contact.get("nextstep") or "",
            "notes": " | ".join(
                [x for x in [str(get(r, "pitch") or "").strip(), str(get(r, "fulfil") or "").strip()] if x]
            ),
            "priority_tier": priority,
            "action": action,
            "demand_low": as_int(get(r, "demandLow"), 0),
            "demand_high": as_int(get(r, "demandHigh"), 0),
            "channels": channels,
            "channel_share_note": str(get(r, "platforms") or "").strip(),
            "competitors_serving": [],
            "competitor_wallet_share": "",
            "score_qc_urgency": clamp_rating(get(r, "qc")),
            "score_sku_fit": clamp_rating(get(r, "sku")),
            "score_order_density": clamp_rating(get(r, "density")),
            "score_pilot_willing": clamp_rating(get(r, "pilot")),
            "score_accessibility": clamp_rating(get(r, "access")),
            "score_logo_value": clamp_rating(get(r, "logo")),
            "bin_slots": 0,
            "price_per_bin": 1500,
            "weighted_mrr": 0,
            "fu1_date": None,
            "fu1_contact": contact.get("name") or "",
            "fu1_mode": "Call",
            "fu1_status": "Pending",
            "fu1_note": "",
            "fu2_date": None,
            "fu2_contact": "",
            "fu2_mode": "Email",
            "fu2_status": "Pending",
            "fu2_note": "",
            "next_followup_date": None,
            "commercial_ask": "",
            "risks": "",
            "score_basis": "Excel import baseline",
            "last_contact_at": None,
            "created_at": parse_datetime(get(r, "createdAt")),
            "updated_at": parse_datetime(get(r, "updatedAt")),
        }

        row["qc_score"] = calc_qc_score(row)
        rows.append(row)

    summary = {
        "rows": len(rows),
        "contact_mapped": sum(1 for r in rows if r.get("contact_name")),
        "with_channels": sum(1 for r in rows if r.get("channels")),
        "with_scores": sum(1 for r in rows if r.get("qc_score", 0) > 0),
    }
    return rows, summary


def post_upsert(url: str, key: str, table: str, rows: List[Dict[str, Any]], on_conflict: str = "legacy_id") -> None:
    endpoint = (
        f"{url.rstrip('/')}/rest/v1/{table}"
        f"?on_conflict={parse.quote(on_conflict)}"
    )

    data = json.dumps(rows).encode("utf-8")
    req = request.Request(endpoint, data=data, method="POST")
    req.add_header("apikey", key)
    req.add_header("Authorization", f"Bearer {key}")
    req.add_header("Content-Type", "application/json")
    req.add_header("Prefer", "resolution=merge-duplicates,return=minimal")

    try:
        with request.urlopen(req, timeout=60) as resp:
            if resp.status not in (200, 201, 204):
                body = resp.read().decode("utf-8", errors="replace")
                raise RuntimeError(f"Supabase upsert failed ({resp.status}): {body}")
    except error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Supabase upsert failed ({exc.code}): {body}") from exc


def delete_all_rows(url: str, key: str, table: str) -> None:
    endpoint = f"{url.rstrip('/')}/rest/v1/{table}?id=not.is.null"
    req = request.Request(endpoint, method="DELETE")
    req.add_header("apikey", key)
    req.add_header("Authorization", f"Bearer {key}")
    req.add_header("Prefer", "return=minimal")

    try:
        with request.urlopen(req, timeout=60) as resp:
            if resp.status not in (200, 204):
                body = resp.read().decode("utf-8", errors="replace")
                raise RuntimeError(f"Supabase delete failed ({resp.status}): {body}")
    except error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Supabase delete failed ({exc.code}): {body}") from exc


def chunked(items: List[Dict[str, Any]], size: int) -> List[List[Dict[str, Any]]]:
    return [items[i : i + size] for i in range(0, len(items), size)]


def main() -> int:
    args = cli()

    if not os.path.exists(args.excel):
        print(f"Excel file not found: {args.excel}", file=sys.stderr)
        return 2

    url = os.getenv("SUPABASE_URL", "").strip()
    key = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "").strip() or os.getenv("SUPABASE_ANON_KEY", "").strip()

    rows, summary = build_accounts_payload(args.excel)
    channel_rows = channel_rows_from_accounts(rows)
    print("Prepared payload:")
    print(json.dumps(summary, indent=2))

    if args.dry_run:
        print("Dry-run mode: no rows were sent.")
        return 0

    if not url or not key:
        print("Missing SUPABASE_URL and/or SUPABASE_SERVICE_ROLE_KEY (or SUPABASE_ANON_KEY).", file=sys.stderr)
        return 2

    if args.reset:
        delete_all_rows(url, key, args.table)
        print(f"Reset complete: {args.table}")

    if args.reset_competitors:
        delete_all_rows(url, key, args.competitors_table)
        print(f"Reset complete: {args.competitors_table}")

    if channel_rows:
        post_upsert(url, key, args.channels_table, channel_rows, on_conflict="name")
        print(f"Upserted {len(channel_rows)} channels into {args.channels_table}")

    batches = chunked(rows, max(1, args.batch_size))
    for idx, batch in enumerate(batches, start=1):
        post_upsert(url, key, args.table, batch)
        print(f"Upserted batch {idx}/{len(batches)} ({len(batch)} rows)")

    print(f"Done. Upserted {len(rows)} rows into {args.table}.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
